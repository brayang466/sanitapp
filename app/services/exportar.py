from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.registro_formato import RegistroFormato


COLUMNAS = [
    "FECHA",
    "ÁREA",
    "EQUIPO O SUPERFICIE",
    "FRECUENCIA",
    "NO CUMPLE",
    "CUMPLE",
    "MEDIDA CORRECTIVA",
    "RESPONSABLE",
]


def _filas(registros: list[RegistroFormato]) -> list[list]:
    filas = []
    for r in registros:
        filas.append(
            [
                r.fecha.strftime("%d/%m/%Y"),
                r.area,
                r.equipo_o_superficie,
                r.frecuencia,
                "X" if r.no_cumple else "",
                "X" if r.cumple else "",
                r.medida_correctiva or "",
                r.responsable,
            ]
        )
    return filas


def exportar_excel(registros: list[RegistroFormato], titulo: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0D9488")
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, nombre in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=3, column=col, value=nombre)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="115E59")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, fila in enumerate(_filas(registros), 4):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    from openpyxl.utils import get_column_letter

    anchos = [12, 28, 38, 14, 10, 10, 32, 22]
    for idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_pdf(registros: list[RegistroFormato], titulo: str) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=28,
        rightMargin=28,
        topMargin=36,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"<b>{titulo}</b>", styles["Title"]),
        Spacer(1, 12),
    ]

    datos = [COLUMNAS] + _filas(registros)
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdfa")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return buffer
